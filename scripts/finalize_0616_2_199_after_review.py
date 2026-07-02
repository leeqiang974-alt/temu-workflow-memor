from __future__ import annotations

import argparse
import hashlib
import html
import json
import re
import sys
import time
import uuid
from collections import Counter, defaultdict
from pathlib import Path

import oss2
from openpyxl import load_workbook
from PIL import Image


COMFYUI_BASE = Path(r"C:\Users\Administrator\Documents\Codex\2026-06-08\comfyui")
if str(COMFYUI_BASE / "work") not in sys.path:
    sys.path.insert(0, str(COMFYUI_BASE / "work"))

import generate_apply_ali_tfirst_0608 as ali_t


DXX_BASE = Path(r"D:\Desktop\jit\DXXmall")
SOURCE_199 = DXX_BASE / "0616-2_TRUE_FINAL_202提交前_尺寸T4最终校验_剔除已复制D_20260701_162622_完整流程_复检前_20260701_162833.xlsx"
OUT_DIR = DXX_BASE / "outputs" / "store_newskill_final_199_writeback_20260702"
INTERMEDIATE_WORKBOOK = OUT_DIR / "0616-2_199_最终回传_T标题已回写_J待重做_20260702.xlsx"
FINAL_WORKBOOK = OUT_DIR / "0616-2_199_最终回传_T标题J全量重做_20260702.xlsx"
VALIDATION_JSON = OUT_DIR / "final_validation.json"
VALIDATION_HTML = OUT_DIR / "final_validation.html"
T_MANIFEST = OUT_DIR / "t_upload_manifest.json"
T_COMPRESSED_DIR = OUT_DIR / "t_first_800"

FULL_RECORDS = DXX_BASE / "outputs" / "store_newskill_image2_full_0616_2_20260701" / "0616_2_image2_199_t_review_records.json"
REDO_RECORDS = DXX_BASE / "outputs" / "store_newskill_image2_redo_22_20260702" / "0616_2_image2_redo_22_review_records.json"
SEEDREAM_9_RECORDS = DXX_BASE / "outputs" / "store_newskill_seedream_fallback_9_20260702" / "0616_2_seedream_fallback_9_review_records.json"
SEEDREAM_1_RECORDS = DXX_BASE / "outputs" / "store_newskill_seedream_fallback_9_20260702" / "0616_2_seedream_fallback_1_L043060505_review_records.json"

OSS_PREFIX = "temu-jit/dxxmall-0616-2/final-199-t-first-newskill"
TARGET_SIZE = 800
TARGET_MAX_BYTES = 150 * 1024

IMAGE2_REDO_PASS = {
    "L043060502",
    "L047060502",
    "L047060505",
    "L047060507",
    "L047060508",
    "L047060509",
    "L068060503",
    "L068060504",
    "L082060502",
    "L082060506",
    "L082060508",
    "L085060504",
}

SEEDREAM_PASS = {
    "L042060502",
    "L042060505",
    "L043060501",
    "L043060503",
    "L043060504",
    "L043060505",
    "L043060506",
    "L043060507",
    "L043060508",
    "L082060501",
}

FORBIDDEN_TITLE_WORDS = [
    "自动",
    "智能",
    "AI",
    "APP",
    "蓝牙",
    "wifi",
    "遥控",
    "感应",
    "电动",
    "充电",
    "电池",
    "USB",
    "儿童",
    "玩具",
    "环保",
    "有机",
    "可降解",
    "食品级",
    "最佳",
    "最好",
    "顶级",
    "久用不塌陷",
    "长久不变形",
    "亲肤",
    "高弹填充",
    "四面延伸",
]


def split_urls(value: object) -> list[str]:
    return [part.strip() for part in re.split(r"[\r\n]+", str(value or "")) if part.strip()]


def dedupe(values: list[str]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        if value and value not in seen:
            result.append(value)
            seen.add(value)
    return result


def stable_tracking_code(d_value: str, used: set[str]) -> str:
    letters = "ABCDEFGHJKLMNPQRSTUVWXYZ"
    digits = "23456789"
    nonce = 0
    while True:
        digest = hashlib.sha256(f"0616-2-final-199-title-fingerprint|{d_value}|{nonce}".encode("utf-8")).digest()
        code = letters[digest[0] % len(letters)] + digits[digest[1] % len(digits)] + letters[digest[2] % len(letters)]
        if code not in used:
            used.add(code)
            return code
        nonce += 1


def apply_tracking_code(title: object, code: str) -> str:
    text = re.sub(r"\s+[A-Z][0-9][A-Z]\s*$", "", str(title or "").strip())
    if len(text) + 4 > 80:
        text = text[:76].rstrip("，、 -")
    return f"{text} {code}".strip()


def headers(ws) -> list[str]:
    return [ws.cell(1, column).value for column in range(1, ws.max_column + 1)]


def col(header_values: list[str], name: str) -> int:
    return header_values.index(name) + 1


def load_json(path: Path):
    return json.loads(path.read_text(encoding="utf-8-sig"))


def records_by_d(path: Path) -> dict[str, dict]:
    records = load_json(path)
    result: dict[str, dict] = {}
    for record in records:
        d_value = str(record.get("d") or record.get("D") or record.get("d_value") or "").strip()
        if d_value:
            result[d_value] = record
    return result


def record_local_path(record: dict) -> Path:
    for key in ("local_path", "local_image", "output_path"):
        value = record.get(key)
        if value:
            return Path(str(value))
    raise KeyError(f"record has no local path: {record}")


def final_t_records() -> tuple[dict[str, dict], dict[str, int]]:
    full = records_by_d(FULL_RECORDS)
    redo = records_by_d(REDO_RECORDS)
    seedream = records_by_d(SEEDREAM_9_RECORDS)
    seedream.update(records_by_d(SEEDREAM_1_RECORDS))

    final: dict[str, dict] = {}
    for d_value, record in full.items():
        final[d_value] = {**record, "selected_batch": "image2_full"}
    for d_value in IMAGE2_REDO_PASS:
        if d_value not in redo:
            raise RuntimeError(f"missing image2 redo pass record: {d_value}")
        final[d_value] = {**redo[d_value], "selected_batch": "image2_redo_pass"}
    for d_value in SEEDREAM_PASS:
        if d_value not in seedream:
            raise RuntimeError(f"missing seedream pass record: {d_value}")
        final[d_value] = {**seedream[d_value], "selected_batch": "seedream_fallback_pass"}

    for d_value, record in final.items():
        path = record_local_path(record)
        if not path.exists():
            raise FileNotFoundError(f"{d_value}: {path}")
    return final, dict(Counter(record["selected_batch"] for record in final.values()))


def resize_compress(input_path: Path, output_path: Path) -> dict:
    image = Image.open(input_path).convert("RGB").resize((TARGET_SIZE, TARGET_SIZE), Image.Resampling.LANCZOS)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    from io import BytesIO

    best = b""
    best_quality = 88
    for quality in range(88, 34, -3):
        buffer = BytesIO()
        image.save(buffer, format="JPEG", quality=quality, optimize=True, progressive=True, subsampling=1)
        best = buffer.getvalue()
        best_quality = quality
        if len(best) <= TARGET_MAX_BYTES:
            break
    output_path.write_bytes(best)
    return {"width": TARGET_SIZE, "height": TARGET_SIZE, "bytes": len(best), "quality": best_quality}


def load_manifest() -> dict:
    if T_MANIFEST.exists():
        return load_json(T_MANIFEST)
    return {"uploaded": {}}


def save_manifest(manifest: dict) -> None:
    T_MANIFEST.parent.mkdir(parents=True, exist_ok=True)
    T_MANIFEST.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")


def upload_once(bucket: oss2.Bucket, manifest: dict, d_value: str, path: Path) -> str:
    key = f"{d_value}|{path.resolve()}"
    uploaded = manifest.setdefault("uploaded", {})
    if key in uploaded:
        return uploaded[key]
    object_key = f"{OSS_PREFIX}/{time.strftime('%Y%m%d')}/{uuid.uuid4().hex}_{d_value}_tfirst_800.jpg"
    bucket.put_object_from_file(object_key, str(path), headers={"Content-Type": "image/jpeg"})
    url = f"https://{bucket.bucket_name}.{bucket.endpoint.replace('https://', '').replace('http://', '')}/{object_key}"
    uploaded[key] = url
    save_manifest(manifest)
    return url


def build_t_urls(new_first: str, old_urls: list[str]) -> list[str]:
    size_url = old_urls[3] if len(old_urls) >= 4 else ""
    rest = [url for url in old_urls if url != new_first and url != size_url]
    if size_url:
        urls = [new_first] + rest[:2] + [size_url] + rest[2:]
    else:
        urls = [new_first] + rest
    return dedupe(urls)[:10]


def prepare_t_workbook() -> int:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    records, batch_counts = final_t_records()

    oss_config = ali_t.ali.read_oss_config()
    bucket = oss2.Bucket(
        oss2.Auth(oss_config["access_key_id"], oss_config["access_key_secret"]),
        f"https://{oss_config['endpoint']}",
        oss_config["bucket"],
    )
    manifest = load_manifest()
    uploaded_by_d: dict[str, str] = {}
    compress_report: dict[str, dict] = {}
    selected_records: dict[str, dict] = {}
    for d_value, record in sorted(records.items()):
        local_path = record_local_path(record)
        compressed = T_COMPRESSED_DIR / d_value[:4] / f"{d_value}_final_199_tfirst_800.jpg"
        compress_report[d_value] = resize_compress(local_path, compressed)
        uploaded_by_d[d_value] = upload_once(bucket, manifest, d_value, compressed)
        selected_records[d_value] = {
            "selected_batch": record.get("selected_batch"),
            "source_id": record.get("source_id"),
            "source_png": record.get("source_png"),
            "local_path": str(local_path),
            "compressed": str(compressed),
            "oss_url": uploaded_by_d[d_value],
        }

    wb = load_workbook(SOURCE_199)
    ws = wb.active
    h = headers(ws)
    col_title = col(h, "产品标题")
    col_d = col(h, "产品货号")
    col_t = col(h, "轮播图")
    col_u = col(h, "产品素材图")

    effective_ds: set[str] = set()
    original_fourth_by_d: dict[str, str] = {}
    changed_rows = []
    title_codes: dict[str, str] = {}
    used_title_codes: set[str] = set()
    for row in range(2, ws.max_row + 1):
        d_value = str(ws.cell(row, col_d).value or "").strip()
        if not d_value:
            continue
        effective_ds.add(d_value)
        title_codes.setdefault(d_value, stable_tracking_code(d_value, used_title_codes))
        ws.cell(row, col_title).value = apply_tracking_code(ws.cell(row, col_title).value, title_codes[d_value])
        old_urls = split_urls(ws.cell(row, col_t).value)
        if len(old_urls) >= 4:
            original_fourth_by_d.setdefault(d_value, old_urls[3])
        if d_value not in uploaded_by_d:
            continue
        final_urls = build_t_urls(uploaded_by_d[d_value], old_urls)
        ws.cell(row, col_t).value = "\n".join(final_urls)
        ws.cell(row, col_u).value = final_urls[0]
        changed_rows.append({"row": row, "d": d_value, "t_count": len(final_urls)})

    missing = sorted(effective_ds - set(uploaded_by_d))
    extra = sorted(set(uploaded_by_d) - effective_ds)
    if missing or extra:
        raise RuntimeError(f"T map mismatch missing={missing[:20]} extra={extra[:20]}")

    wb.save(INTERMEDIATE_WORKBOOK)
    wb.close()

    report = {
        "source_workbook": str(SOURCE_199),
        "title_fingerprint_rule": "Use this 199 workbook's current rewritten title as base; append one unique deterministic tracking code per D for this file. Do not copy title fingerprints from the 202 source workbook.",
        "intermediate_workbook": str(INTERMEDIATE_WORKBOOK),
        "final_workbook_target": str(FINAL_WORKBOOK),
        "batch_counts": batch_counts,
        "effective_unique_d": len(effective_ds),
        "selected_unique_d": len(uploaded_by_d),
        "changed_rows_count": len(changed_rows),
        "changed_rows": changed_rows,
        "title_codes": title_codes,
        "original_fourth_by_d": original_fourth_by_d,
        "selected_records": selected_records,
        "compression": {
            "count": len(compress_report),
            "max_bytes": max(item["bytes"] for item in compress_report.values()),
            "over_target": {d: item for d, item in compress_report.items() if item["bytes"] > TARGET_MAX_BYTES},
        },
    }
    (OUT_DIR / "t_prepare_report.json").write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({k: v for k, v in report.items() if k not in {"changed_rows", "original_fourth_by_d", "selected_records"}}, ensure_ascii=False, indent=2))
    return 0


def load_values_by_key(path: Path, key_cols: list[str]) -> dict[tuple, dict[str, object]]:
    wb = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    ws = wb.active
    h = headers(ws)
    cols = {name: col(h, name) for name in h if name}
    key_idx = [cols[name] for name in key_cols]
    rows: dict[tuple, dict[str, object]] = {}
    for row in range(2, ws.max_row + 1):
        key = tuple(str(ws.cell(row, idx).value or "").strip() for idx in key_idx)
        if not key[0]:
            continue
        rows[key] = {str(header): ws.cell(row, idx).value for idx, header in enumerate(h, 1) if header}
    wb.close()
    return rows


def validate_final() -> int:
    prep = load_json(OUT_DIR / "t_prepare_report.json")
    original_fourth_by_d = prep["original_fourth_by_d"]
    selected_records = prep["selected_records"]
    source_rows = load_values_by_key(SOURCE_199, ["产品货号", "SKU货号", "变种属性值一"])
    final_rows = load_values_by_key(FINAL_WORKBOOK, ["产品货号", "SKU货号", "变种属性值一"])

    wb = load_workbook(FINAL_WORKBOOK, read_only=True, data_only=True, keep_links=False)
    ws = wb.active
    h = headers(ws)
    c = {name: col(h, name) for name in h if name}
    rows_by_d: dict[str, list[int]] = defaultdict(list)
    title_by_d: dict[str, str] = {}
    code_by_d: dict[str, str] = {}
    t_by_d: dict[str, str] = {}
    u_by_d: dict[str, str] = {}
    errors: dict[str, list] = defaultdict(list)
    j_records = load_json(OUT_DIR / "j_rebuild_report.json") if (OUT_DIR / "j_rebuild_report.json").exists() else {}
    j_rows = {record.get("row"): record for record in j_records.get("records", []) if record.get("status") == "ok"}

    for row in range(2, ws.max_row + 1):
        d_value = str(ws.cell(row, c["产品货号"]).value or "").strip()
        if not d_value:
            continue
        rows_by_d[d_value].append(row)
        title = str(ws.cell(row, c["产品标题"]).value or "").strip()
        j_url = str(ws.cell(row, c["预览图"]).value or "").strip()
        t_urls = split_urls(ws.cell(row, c["轮播图"]).value)
        u_url = str(ws.cell(row, c["产品素材图"]).value or "").strip()
        key = (
            d_value,
            str(ws.cell(row, c["SKU货号"]).value or "").strip(),
            str(ws.cell(row, c["变种属性值一"]).value or "").strip(),
        )

        if not re.search(r"\s[A-Z][0-9][A-Z]$", title):
            errors["missing_tracking_code"].append({"row": row, "d": d_value, "title": title})
        else:
            code_by_d.setdefault(d_value, title.rsplit(" ", 1)[-1])
        bad_words = [word for word in FORBIDDEN_TITLE_WORDS if word.lower() in title.lower()]
        if bad_words:
            errors["forbidden_title_words"].append({"row": row, "d": d_value, "words": bad_words, "title": title})
        if title_by_d.get(d_value, title) != title:
            errors["same_d_title_mismatch"].append({"row": row, "d": d_value})
        title_by_d.setdefault(d_value, title)
        if t_by_d.get(d_value, "\n".join(t_urls)) != "\n".join(t_urls):
            errors["same_d_t_mismatch"].append({"row": row, "d": d_value})
        t_by_d.setdefault(d_value, "\n".join(t_urls))
        if u_by_d.get(d_value, u_url) != u_url:
            errors["same_d_u_mismatch"].append({"row": row, "d": d_value})
        u_by_d.setdefault(d_value, u_url)

        if not j_url:
            errors["empty_j"].append({"row": row, "d": d_value})
        if not t_urls:
            errors["empty_t"].append({"row": row, "d": d_value})
        if len(t_urls) > 10:
            errors["t_over_10"].append({"row": row, "d": d_value, "count": len(t_urls)})
        if t_urls and u_url != t_urls[0]:
            errors["u_not_t_first"].append({"row": row, "d": d_value})
        if d_value in selected_records and t_urls and t_urls[0] != selected_records[d_value]["oss_url"]:
            errors["t_first_not_selected"].append({"row": row, "d": d_value})
        expected_fourth = original_fourth_by_d.get(d_value, "")
        if expected_fourth and (len(t_urls) < 4 or t_urls[3] != expected_fourth):
            errors["fourth_image_changed"].append({"row": row, "d": d_value})
        if row in j_rows and j_url != j_rows[row].get("oss_url"):
            errors["j_url_mismatch_report"].append({"row": row, "d": d_value})
        if key not in source_rows:
            errors["row_key_missing_in_source"].append({"row": row, "key": key})
        else:
            src = source_rows[key]
            dst = final_rows.get(key, {})
            for name in ("SPUID", "SKCID", "SKUID", "分类id", "SKU货号"):
                if str(src.get(name) or "") != str(dst.get(name) or ""):
                    errors["identity_column_changed"].append({"row": row, "d": d_value, "column": name, "source": src.get(name), "final": dst.get(name)})

    wb.close()
    duplicate_codes = [code for code, count in Counter(code_by_d.values()).items() if count > 1]
    if duplicate_codes:
        errors["duplicate_tracking_code"].extend([{"code": code} for code in duplicate_codes])
    missing_selected = sorted(set(rows_by_d) - set(selected_records))
    extra_selected = sorted(set(selected_records) - set(rows_by_d))
    error_counts = {name: len(items) for name, items in sorted(errors.items())}
    validation = {
        "final_workbook": str(FINAL_WORKBOOK),
        "effective_rows": sum(len(rows) for rows in rows_by_d.values()),
        "unique_d": len(rows_by_d),
        "selected_t_unique_d": len(selected_records),
        "missing_selected_d": missing_selected,
        "extra_selected_d": extra_selected,
        "j_records_ok": len(j_rows),
        "j_match_modes": dict(Counter(record.get("match_mode") for record in j_rows.values())),
        "error_counts": error_counts,
        "error_count": sum(error_counts.values()) + len(missing_selected) + len(extra_selected),
        "errors": {name: items[:50] for name, items in sorted(errors.items())},
    }
    VALIDATION_JSON.write_text(json.dumps(validation, ensure_ascii=False, indent=2), encoding="utf-8")
    VALIDATION_HTML.write_text(
        "<!doctype html><meta charset='utf-8'><title>0616-2 199 final validation</title>"
        "<style>body{font-family:Arial,'Microsoft YaHei',sans-serif;margin:24px;background:#f7f4ef}"
        ".box{background:white;border:1px solid #ddd;border-radius:8px;padding:12px;margin:12px 0}"
        "pre{white-space:pre-wrap;background:#f6f6f6;padding:10px}</style>"
        f"<h1>0616-2 199 final validation</h1><div class='box'>"
        f"<p>Final: <code>{html.escape(str(FINAL_WORKBOOK))}</code></p>"
        f"<p>Rows: {validation['effective_rows']} Unique D: {validation['unique_d']} Error count: <b>{validation['error_count']}</b></p>"
        f"</div><div class='box'><h2>Error counts</h2><pre>{html.escape(json.dumps(error_counts, ensure_ascii=False, indent=2))}</pre></div>"
        f"<div class='box'><h2>J match modes</h2><pre>{html.escape(json.dumps(validation['j_match_modes'], ensure_ascii=False, indent=2))}</pre></div>"
        f"<div class='box'><h2>Examples</h2><pre>{html.escape(json.dumps(validation['errors'], ensure_ascii=False, indent=2))}</pre></div>",
        encoding="utf-8",
    )
    print(json.dumps({k: v for k, v in validation.items() if k not in {"errors"}}, ensure_ascii=False, indent=2))
    return 0 if validation["error_count"] == 0 else 1


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--mode", choices=["prepare-t", "validate-final", "paths"], required=True)
    args = parser.parse_args()
    if args.mode == "prepare-t":
        return prepare_t_workbook()
    if args.mode == "validate-final":
        return validate_final()
    print(json.dumps({
        "source_199": str(SOURCE_199),
        "title_fingerprint_rule": "per-file/per-D unique code generated from current 199 workbook titles",
        "out_dir": str(OUT_DIR),
        "intermediate_workbook": str(INTERMEDIATE_WORKBOOK),
        "final_workbook": str(FINAL_WORKBOOK),
        "validation_json": str(VALIDATION_JSON),
        "validation_html": str(VALIDATION_HTML),
    }, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
