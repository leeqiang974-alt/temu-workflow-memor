from __future__ import annotations

import argparse
import html
import json
import math
import random
import re
from datetime import datetime
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageEnhance, ImageFilter, ImageOps


IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".webp"}
BAD_WORDS = {"九宫格", "9grid", "out", "output", "背景素材"}
SIZE = 800

TEXTURE_STYLES = [
    ("sand_terracotta_plaster", [(246, 226, 190), (229, 176, 142), (248, 238, 206)], "plaster"),
    ("sage_yellow_green_stone", [(203, 228, 194), (239, 228, 171), (178, 211, 185)], "stone"),
    ("sky_blue_cream_mineral", [(194, 222, 238), (232, 221, 196), (177, 200, 218)], "mineral"),
    ("rose_peach_clay_plaster", [(245, 198, 184), (222, 160, 144), (250, 225, 198)], "plaster"),
    ("oat_beige_linen", [(246, 233, 203), (214, 190, 158), (250, 243, 218)], "linen"),
    ("aqua_mint_ceramic", [(191, 230, 224), (164, 205, 220), (239, 226, 188)], "ceramic"),
    ("lavender_greige_mineral", [(224, 211, 232), (207, 196, 182), (238, 229, 216)], "mineral"),
    ("butter_blue_limewash", [(245, 229, 164), (195, 221, 232), (246, 237, 204)], "plaster"),
]

TOKEN_MAP = [
    ("black", ["黑", "black"]),
    ("green", ["绿", "green"]),
    ("white", ["白", "透明", "white"]),
    ("gray", ["灰", "grey", "gray"]),
    ("yellow", ["黄", "yellow"]),
    ("pink", ["粉", "pink"]),
    ("red", ["红", "red"]),
    ("blue", ["蓝", "blue"]),
    ("purple", ["紫", "purple"]),
    ("beige", ["杏", "米", "beige"]),
    ("wood", ["原木", "木色", "胡桃", "wood"]),
    ("15", ["15"]),
    ("30", ["30"]),
    ("50", ["50"]),
    ("2", ["2", "双", "两", "二", "2格", "2层"]),
    ("3", ["3", "三", "3格", "3层"]),
]


def natural_key(text: object) -> list[object]:
    return [int(part) if part.isdigit() else part.lower() for part in re.split(r"(\d+)", str(text))]


def stable_int(text: str) -> int:
    import hashlib

    return int(hashlib.sha256(text.encode("utf-8")).hexdigest()[:16], 16)


def tokens(text: object) -> set[str]:
    value = str(text or "").lower()
    return {key for key, words in TOKEN_MAP if any(word in value for word in words)}


def clean_image(path: Path) -> bool:
    if not path.is_file() or path.suffix.lower() not in IMAGE_EXTS:
        return False
    lower = str(path).lower()
    return not any(word.lower() in lower for word in BAD_WORDS)


def first_level_files(folder: Path) -> list[Path]:
    if not folder.exists():
        return []
    return sorted([p for p in folder.iterdir() if clean_image(p)], key=lambda p: natural_key(p.name))


def find_sku_folder(prefix: str, sku_roots: list[Path], folder_names: list[str]) -> Path | None:
    for root in sku_roots:
        for folder_name in folder_names:
            folder = root / prefix / folder_name
            if folder.exists() and (first_level_files(folder) or any(child.is_dir() for child in folder.iterdir())):
                return folder
    return None


def strict_match(prefix: str, d_value: str, variant: str, sku: str, sku_roots: list[Path], folder_names: list[str]) -> dict:
    folder = find_sku_folder(prefix, sku_roots, folder_names)
    wanted_tokens = tokens(" ".join([variant, sku]))
    if folder is None:
        return {"status": "missing", "wanted_tokens": sorted(wanted_tokens), "sku_root": " | ".join(str(root / prefix) for root in sku_roots)}

    candidates: list[dict] = []
    direct_files = first_level_files(folder)
    if direct_files:
        for file in direct_files:
            got = tokens(file.name)
            candidates.append({
                "mode": "direct_file",
                "path": file,
                "score": len(wanted_tokens & got) * 100 + (10 if sku and sku.lower() in file.name.lower() else 0),
                "tokens": sorted(got),
            })
    else:
        for child in sorted([p for p in folder.iterdir() if p.is_dir()], key=lambda p: natural_key(p.name)):
            if any(word.lower() in str(child).lower() for word in BAD_WORDS):
                continue
            files = first_level_files(child)
            if not files:
                continue
            folder_tokens = tokens(child.name)
            chosen_file = files[stable_int(f"{d_value}|{variant}|{sku}|{child.name}") % len(files)]
            candidates.append({
                "mode": "variant_folder",
                "path": chosen_file,
                "folder": child,
                "score": len(wanted_tokens & folder_tokens) * 100 + (10 if sku and sku.lower() in child.name.lower() else 0),
                "tokens": sorted(folder_tokens),
            })

    if not candidates:
        return {"status": "missing", "wanted_tokens": sorted(wanted_tokens), "sku_root": str(folder)}

    candidates.sort(key=lambda item: (-item["score"], natural_key(str(item.get("folder") or item["path"]))))
    best = candidates[0]
    warning = "no_positive_variant_token_match" if best["score"] <= 0 and wanted_tokens else ""
    return {
        "status": "ok",
        "wanted_tokens": sorted(wanted_tokens),
        "mode": best["mode"],
        "score": best["score"],
        "matched_tokens": best["tokens"],
        "sku_source": str(best["path"]),
        "sku_root": str(folder),
        "variant_folder": str(best.get("folder", "")),
        "warning": warning,
        "candidate_count": len(candidates),
    }


def crop_alpha(path: Path) -> Image.Image:
    image = ImageOps.exif_transpose(Image.open(path)).convert("RGBA")
    bbox = image.getbbox()
    return image.crop(bbox) if bbox else image


def make_bg(style_index: int, seed: int) -> Image.Image:
    _, colors, kind = TEXTURE_STYLES[style_index % len(TEXTURE_STYLES)]
    rng = random.Random(seed % 2147483647)
    image = Image.new("RGB", (SIZE, SIZE), colors[0])
    pixels = image.load()
    angle = rng.uniform(-0.55, 0.55)
    center_x = rng.uniform(0.18, 0.82) * SIZE
    center_y = rng.uniform(0.15, 0.82) * SIZE
    for y in range(SIZE):
        for x in range(SIZE):
            linear = (y / SIZE) * 0.68 + (x / SIZE) * angle * 0.16
            radial = math.hypot((x - center_x) / SIZE, (y - center_y) / SIZE) * 0.28
            wave = math.sin(x * 0.004 + y * 0.004 + (seed % 31)) * 0.025
            ratio = max(0, min(0.999, linear + radial + wave))
            position = ratio * (len(colors) - 1)
            color_index = min(len(colors) - 2, int(position))
            local_ratio = position - color_index
            color_a = colors[color_index]
            color_b = colors[color_index + 1]
            pixels[x, y] = tuple(int(color_a[i] * (1 - local_ratio) + color_b[i] * local_ratio) for i in range(3))

    overlay = Image.new("RGBA", (SIZE, SIZE), (0, 0, 0, 0))
    draw = ImageDraw.Draw(overlay, "RGBA")
    for _ in range(22):
        x = rng.randint(-160, SIZE + 80)
        y = rng.randint(-140, SIZE + 80)
        w = rng.randint(240, 620)
        h = rng.randint(70, 210)
        draw.ellipse((x, y, x + w, y + h), outline=(255, 255, 255, rng.randint(8, 22)), width=1)
    if kind == "linen":
        for _ in range(90):
            x = rng.randint(0, SIZE)
            y = rng.randint(0, SIZE)
            draw.line((x, y, min(SIZE, x + rng.randint(35, 150)), y + rng.randint(-1, 1)), fill=(255, 255, 255, rng.randint(8, 18)), width=1)
    image = Image.alpha_composite(image.convert("RGBA"), overlay).convert("RGB")
    return ImageEnhance.Contrast(ImageEnhance.Color(image).enhance(1.04)).enhance(1.06).convert("RGBA")


def paste_shadow(canvas: Image.Image, image: Image.Image, x: int, y: int, seed: int) -> None:
    alpha = image.getchannel("A")
    shadow = Image.new("RGBA", image.size, (0, 0, 0, 0))
    shadow.putalpha(alpha.filter(ImageFilter.GaussianBlur(9 + seed % 5)))
    canvas.paste(shadow, (x + 8, y + 10), shadow)
    canvas.paste(image, (x, y), image)


def save_j_preview(sku_path: Path, out_path: Path, seed: int, scale_factor: float) -> dict:
    style_index = seed % len(TEXTURE_STYLES)
    style_name = TEXTURE_STYLES[style_index][0]
    canvas = make_bg(style_index, seed)
    slots = [(0, 0, 400, 270), (400, 0, 400, 270), (200, 265, 400, 270), (0, 530, 400, 270), (400, 530, 400, 270)]
    for idx, (left, top, width, height) in enumerate(slots):
        sku = crop_alpha(sku_path)
        scale = min((width * scale_factor) / sku.width, (height * scale_factor) / sku.height, 1.55)
        sku = sku.resize((max(1, int(sku.width * scale)), max(1, int(sku.height * scale))), Image.Resampling.LANCZOS)
        x = left + (width - sku.width) // 2 + [-8, 8, 0, -6, 6][idx]
        y = top + (height - sku.height) // 2 + [5, -5, 0, -4, 4][idx]
        paste_shadow(canvas, sku, x, y, seed + idx * 29)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    rgb = canvas.convert("RGB")
    best = None
    quality_used = 88
    for quality in range(90, 48, -4):
        buffer = BytesIO()
        rgb.save(buffer, format="JPEG", quality=quality, optimize=True, progressive=True, subsampling=1)
        best = buffer.getvalue()
        quality_used = quality
        if len(best) <= 150 * 1024:
            break
    out_path.write_bytes(best or b"")
    return {"width": SIZE, "height": SIZE, "bytes": out_path.stat().st_size, "quality": quality_used, "style": style_name}


def esc(value: object) -> str:
    return html.escape(str(value or ""))


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate row-level J preview images with strict variant matching and five-grid textured backgrounds.")
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--sku-root", required=True, action="append", type=Path, help="Root containing L0xx folders. Can be passed multiple times.")
    parser.add_argument("--out", required=True, type=Path)
    parser.add_argument("--scale", type=float, default=0.72)
    parser.add_argument("--folder-name", action="append", default=["sku文件_最终抠图PNG", "sku文件", "sku"])
    args = parser.parse_args()

    out = args.out
    image_out = out / "images"
    report_json = out / "j_preview_records.json"
    report_html = out / "j_preview_review.html"
    out.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(args.workbook, read_only=True, data_only=True, keep_links=False)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    required = ["产品货号", "变种属性值一", "SKU货号"]
    col = {name: headers.index(name) + 1 for name in required}
    records = []
    for row_index in range(2, ws.max_row + 1):
        d_value = str(ws.cell(row_index, col["产品货号"]).value or "").strip()
        if not d_value:
            continue
        prefix = d_value[:4]
        variant = str(ws.cell(row_index, col["变种属性值一"]).value or "").strip()
        sku = str(ws.cell(row_index, col["SKU货号"]).value or "").strip()
        match = strict_match(prefix, d_value, variant, sku, args.sku_root, args.folder_name)
        record = {"row": row_index, "d": d_value, "prefix": prefix, "variant": variant, "sku": sku, **match}
        if match["status"] == "ok":
            local = image_out / prefix / f"{row_index:04d}_{d_value}_{sku or 'sku'}.jpg"
            info = save_j_preview(Path(match["sku_source"]), local, stable_int(f"{row_index}|{d_value}|{variant}|{sku}"), args.scale)
            record["local_image"] = str(local)
            record["image_info"] = info
            record["style"] = info.get("style", "")
        records.append(record)
    wb.close()

    payload = {
        "generatedAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "workbook": str(args.workbook),
        "record_count": len(records),
        "ok_count": sum(1 for r in records if r["status"] == "ok"),
        "warning_count": sum(1 for r in records if r.get("warning")),
        "missing_count": sum(1 for r in records if r["status"] != "ok"),
        "records": records,
    }
    report_json.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    groups: dict[str, list[dict]] = {}
    for record in records:
        groups.setdefault(record["d"], []).append(record)
    cards = []
    for d_value in sorted(groups, key=natural_key):
        figures = []
        for r in groups[d_value]:
            rel_img = Path(r["local_image"]).relative_to(out).as_posix() if r.get("local_image") else ""
            figures.append(
                f"""<figure class="{esc(r.get('warning') or r['status'])}">
                {f'<img loading="lazy" src="{esc(rel_img)}">' if rel_img else '<div class="missing">缺匹配</div>'}
                <figcaption><b>row {r['row']} · {esc(r['variant'])}</b><br>SKU: {esc(r['sku'])}<br>命中: {esc(r.get('matched_tokens'))} / 需要: {esc(r.get('wanted_tokens'))}<br>源图: {esc(r.get('sku_source'))}<br>{esc(r.get('warning'))}</figcaption>
                </figure>"""
            )
        cards.append(f"<section class='card'><h2>{esc(d_value)} · {esc(d_value[:4])}</h2><div class='grid'>{''.join(figures)}</div></section>")

    report_html.write_text(
        f"""<!doctype html><html lang="zh-CN"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>J 严格变体匹配审核</title><style>
body{{font-family:Arial,'Microsoft YaHei',sans-serif;margin:22px;background:#f7f3ed;color:#241c15}}
.card{{background:#fff;border:1px solid #ead8c5;border-radius:10px;padding:12px;margin:14px 0}}
h1{{font-size:22px}} h2{{font-size:16px}} .grid{{display:grid;grid-template-columns:repeat(4,minmax(180px,1fr));gap:10px}}
figure{{margin:0;border:1px solid #eadfd2;border-radius:8px;padding:7px;background:#fffdf8}} figure.no_positive_variant_token_match{{border-color:#d69731;background:#fff8e8}} figure.missing{{border-color:#c44;background:#fff3f3}}
img{{width:100%;height:180px;object-fit:contain;background:#eee;border-radius:6px}} figcaption{{font-size:11px;line-height:1.35;word-break:break-all}} .missing{{height:180px;display:flex;align-items:center;justify-content:center;background:#fff0f0;border-radius:6px}}
@media(max-width:980px){{.grid{{grid-template-columns:repeat(2,1fr)}}}}
</style></head><body><h1>J 严格变体匹配审核</h1><p>共 {payload['record_count']} 行，OK {payload['ok_count']}，需人工关注 {payload['warning_count']}，缺匹配 {payload['missing_count']}。本页只生成本地审核图，尚未写回 Excel/OSS。</p>{''.join(cards)}</body></html>""",
        encoding="utf-8",
    )
    print(json.dumps({"html": str(report_html), "json": str(report_json), "record_count": payload["record_count"], "ok_count": payload["ok_count"], "warning_count": payload["warning_count"], "missing_count": payload["missing_count"]}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
